from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook


class ExcelService:
    @staticmethod
    def get_excel_obj(start_row=1, modifier=None):
        def do_nothing(input_data):
            return input_data

        if modifier is None:
            modifier = do_nothing

        def inner(list_item: list[dict], fields: tuple[tuple[str, str]]):
            workbook = Workbook()
            ws = workbook.active

            for (i, (_key, title)) in enumerate(fields, start=1):
                address = f"{ExcelService.colnum_string(i)}{start_row}"
                ws[address] = title

            for row_index, item in enumerate(list_item, start=start_row + 1):
                for (i, (key, _title)) in enumerate(fields, start=1):
                    address = f"{ExcelService.colnum_string(i)}{row_index}"
                    ws[address] = item.get(key, "")
                    if key == "index":
                        ws[address] = row_index
            modifier(ws)
            return save_virtual_workbook(workbook)

        return inner

    @staticmethod
    def colnum_string(n):
        string = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string
